"""
Amazon ASIN 数据库保存功能
用于保存影片番号与 ASIN 对应关系，方便后续统计和复用
"""

import asyncio
import os
import threading
from datetime import datetime
from pathlib import Path
from typing import TypedDict

from ..utils.file import write_file_atomic


def _save_workbook_atomic(wb, excel_path: Path) -> None:
    """xlsx 原子保存：先写 .tmp 再 os.replace（全库审查 B4）。

    wb.save() 直接原地覆盖时，进程被杀（停止刮削/强关）或磁盘满会让
    zip 半写损坏，下次启动 merge_asin_db_from_backup 的 load_workbook
    异常被吞后整个用户 ASIN 库被静默废弃。同文件 marker 已用
    write_file_atomic，主体数据文件更应对齐。Windows 上 replace 目标被
    占用时抛 PermissionError，与原语义一致向上传。
    """
    tmp_path = excel_path.with_name(excel_path.name + ".tmp")
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, excel_path)
    except Exception:
        # 保存/替换失败时清掉半写 tmp，避免残留占盘（下次保存会重建）
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except OSError:
            pass
        raise


class AsinRecord(TypedDict, total=False):
    """ASIN 记录结构"""

    number: str  # 影片番号
    asin: str  # 亚马逊 ASIN
    product_url: str  # 亚马逊商品详情页链接
    title: str  # 商品标题
    poster_url: str  # 封面图片 URL
    search_keyword: str  # 搜索关键词


# ============================================================
# 查询索引缓存：避免每次查询都全量解析 4 万行 xlsx
# 首次查询建索引（number/asin 双索引），此后命中内存；
# 文件 mtime 变化（写入/合并）后自动失效重建
# ============================================================

_asin_index_lock = threading.Lock()
_asin_index_cache: dict[Path, dict] = {}

# xlsx 写互斥：写路径已挪 asyncio.to_thread，并发刮削协程对同一库文件的
# save 会真并行（load→改→save 交错丢更新）；单写者锁串行化（全库审查 M5）
_asin_db_write_lock = threading.Lock()


def invalidate_asin_cache(excel_path: Path | None = None) -> None:
    """失效查询索引缓存（写入/合并 ASIN 库后调用）；无参清空全部。"""
    with _asin_index_lock:
        if excel_path is None:
            _asin_index_cache.clear()
        else:
            _asin_index_cache.pop(Path(excel_path), None)


def _get_asin_index(excel_path: Path) -> dict | None:
    """取查询索引；文件变化或未建时重建，文件缺失/解析失败返回 None。"""
    import openpyxl

    path = Path(excel_path)
    try:
        mtime = path.stat().st_mtime_ns
    except OSError:
        with _asin_index_lock:
            _asin_index_cache.pop(path, None)
        return None

    with _asin_index_lock:
        cached = _asin_index_cache.get(path)
        if cached is not None and cached["mtime"] == mtime:
            return cached

    # 锁外解析（openpyxl 慢），解析完成后再入缓存并复核 mtime
    number_index: dict[str, list[AsinRecord]] = {}
    asin_index: dict[str, list[AsinRecord]] = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1 or len(row) < 6:
                    continue
                record = AsinRecord(
                    number=str(row[0] or ""),
                    asin=str(row[1] or ""),
                    product_url=str(row[2] or ""),
                    title=str(row[3] or ""),
                    poster_url=str(row[4] or ""),
                    search_keyword=str(row[5] or ""),
                )
                if record["number"]:
                    number_index.setdefault(record["number"].upper(), []).append(record)
                if record["asin"]:
                    asin_index.setdefault(record["asin"].upper(), []).append(record)
        finally:
            wb.close()
    except Exception as e:
        from ..models.log_buffer import LogBuffer

        LogBuffer.log().write(f"  ⚠️ [ASIN 数据库] 读取失败：{e}")
        return None

    try:
        mtime_after = path.stat().st_mtime_ns
    except OSError:
        return None

    index = {"mtime": mtime_after, "number": number_index, "asin": asin_index}
    with _asin_index_lock:
        # 复核期间文件又被改动则放弃本次结果（下次查询重建）
        if mtime_after == mtime:
            _asin_index_cache[path] = index
        else:
            _asin_index_cache.pop(path, None)
    return index


def _get_default_excel_path() -> Path:
    """获取默认的 Excel 文件路径，位于 userdata 目录下（与 mapping、watermark 等同目录）"""
    from ..config.manager import manager

    userdata_dir = manager.data_folder / "userdata"
    userdata_dir.mkdir(parents=True, exist_ok=True)
    return userdata_dir / "amazon_asin_database.xlsx"


def _asin_sort_key(row: tuple) -> tuple:
    """ASIN 库按番号排序键：前缀字母升序 + 数字升序；无法解析的放最后。"""
    import re

    n = str(row[0]) if row and row[0] is not None else ""
    m = re.match(r"^([A-Za-z]+)[-_]?(\d+)", n)
    if not m:
        return (chr(0x10FFFF), n)
    return (m.group(1).upper(), int(m.group(2)))


def _resort_asin_worksheet(local_path: Path) -> None:
    """读取 ASIN 库全部数据行，按番号排序后重建工作簿并重新格式化。

    用重建（而非原地 sort_rows）避免 delete_rows 的 max_row 虚高/空行残留，
    复用 _format_asin_worksheet 保持表头样式/边框/超链接/auto_filter 一致。
    """
    import openpyxl
    from openpyxl import Workbook

    wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = [
        tuple(r[:6])
        for r in ws.iter_rows(min_row=2, max_col=6, values_only=True)
        if r and r[0] is not None and str(r[0]).strip()
    ]
    wb.close()

    rows.sort(key=_asin_sort_key)

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = ws.title
    new_ws.append(header)
    for r in rows:
        new_ws.append(list(r))
    _format_asin_worksheet(new_ws)
    _save_workbook_atomic(new_wb, local_path)
    new_wb.close()


def merge_asin_db_from_backup(backup_path: Path, local_path: Path) -> None:
    """把出厂 ASIN 库同步进用户库（出厂库是权威数据）。

    合并规则：
    - 番号相同：出厂库数据**无条件覆盖**用户库数据（番号外 5 列全替换）。
    - 番号仅在出厂库存在：追加到用户库。
    - 番号仅在用户库存在：保留不删。

    出厂库随软件版本更新（新增/修正番号→ASIN 映射），用户库可能残留历史错配
    或过期数据，以出厂库为权威源整体纠偏。
    合并有变化时，按番号（前缀字母 + 数字）整体重排并重新格式化。

    用出厂库文件 md5 作为合并标记写入 local_path 同目录的 .asin_db_merge_marker，
    出厂库内容未变时跳过，避免每次启动重复扫描。
    """
    from ..models.log_buffer import LogBuffer

    try:
        import importlib.util

        if not importlib.util.find_spec("openpyxl"):
            raise ImportError("openpyxl not found")
    except ImportError:
        LogBuffer.log().write("  ⚠️ [ASIN 数据库] 缺少 openpyxl，无法合并 amazon_asin_database.xlsx")
        return

    if not backup_path.exists() or not local_path.exists():
        return

    import hashlib

    marker_path = local_path.parent / ".asin_db_merge_marker"
    try:
        backup_hash = hashlib.md5(backup_path.read_bytes()).hexdigest()
        if marker_path.exists() and marker_path.read_text(encoding="utf-8").strip() == backup_hash:
            return  # 出厂库未变化，无需合并

        # 与刮削写路径互斥（全库审查 M5）：合并读改写期间并发 save 会丢更新
        with _asin_db_write_lock:
            _merge_asin_db_locked(backup_path, local_path, marker_path, backup_hash, LogBuffer)
    except Exception as e:
        LogBuffer.log().write(f"  ⚠️ [ASIN 数据库] 出厂库合并失败: {e}")


def _merge_asin_db_locked(
    backup_path: Path,
    local_path: Path,
    marker_path: Path,
    backup_hash: str,
    LogBuffer,  # noqa: N803
) -> None:
    import openpyxl  # noqa: F401  # 外壳已验证可导入，此处供本函数体使用

    try:
        wb = openpyxl.load_workbook(local_path)
        ws = wb.active
        number_row_map: dict[str, int] = {}
        next_row = ws.max_row + 1
        for row_no, row in enumerate(ws.iter_rows(min_row=2, max_col=6, values_only=True), start=2):
            if row and row[0]:
                number_row_map.setdefault(str(row[0]).strip().upper(), row_no)

        added = 0
        replaced = 0
        backup_wb = openpyxl.load_workbook(backup_path, read_only=True, data_only=True)
        backup_ws = backup_wb.active
        for row in backup_ws.iter_rows(min_row=2, max_col=6, values_only=True):
            if not row or not row[0]:
                continue
            number = str(row[0]).strip().upper()
            if number in number_row_map:
                # 出厂库权威：番号外 5 列无条件覆盖用户值
                existing_row = number_row_map[number]
                for col_idx in range(1, 6):  # 番号列除外
                    new = row[col_idx] if col_idx < len(row) else None
                    cur = ws.cell(row=existing_row, column=col_idx + 1).value
                    if cur != new:
                        ws.cell(row=existing_row, column=col_idx + 1, value=new)
                        replaced += 1
                continue
            ws.append(list(row[:6]))
            number_row_map[number] = next_row
            next_row += 1
            added += 1
        backup_wb.close()

        if added or replaced:
            _save_workbook_atomic(wb, local_path)
        wb.close()
        if added:
            # 有新增行才整体重排（纯覆盖不改变行数与顺序，无需重排）
            _resort_asin_worksheet(local_path)
        # marker 必须在保存+重排全部成功后写：先写 marker 后重排失败时，
        # 下次启动 marker 匹配跳过合并，库永远停留未排序形态（全库审查发现8）
        write_file_atomic(marker_path, backup_hash, "utf-8")
        invalidate_asin_cache(local_path)
        if added or replaced:
            LogBuffer.log().write(f"  ℹ️ [ASIN 数据库] 出厂库合并: 新增 {added} 条, 覆盖 {replaced} 个字段")
    except Exception as e:
        LogBuffer.log().write(f"  ⚠️ [ASIN 数据库] 出厂库合并失败: {e}")


def _save_asin_to_excel_sync(
    records: list[AsinRecord],
    excel_path: Path,
    sheet_name: str,
) -> None:
    """save_asin_to_excel 的同步核心（load→去重→append→格式化→原子保存）。

    独立成函数供 asyncio.to_thread 调用——openpyxl 全表格式化在 4 万行库上
    耗时秒级，直接在 async 函数体内跑会阻塞整个事件循环，所有并发刮削协程
    （含网络 IO）停摆（全库审查 B4）。
    """

    with _asin_db_write_lock:
        _save_asin_to_excel_locked(records, excel_path, sheet_name)


def _save_asin_to_excel_locked(
    records: list[AsinRecord],
    excel_path: Path,
    sheet_name: str,
) -> None:
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        from ..models.log_buffer import LogBuffer

        LogBuffer.log().write("  ⚠️ [ASIN 数据库] 缺少 openpyxl，无法保存 amazon_asin_database.xlsx")
        raise ImportError("请安装 openpyxl 库：pip install openpyxl") from None

    excel_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    try:
        ws = wb.active
        ws.title = sheet_name

        # 去重：以番号为键，用户库已有该番号时跳过不写（避免重复行）
        existing_numbers: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                existing_numbers.add(str(row[0]).strip().upper())

        if not ws["A1"].value:
            headers = [
                "影片番号",
                "ASIN 编号",
                "影片链接",
                "商品标题",
                "封面 URL",
                "搜索关键词",
            ]
            # 使用 cell() 直接设置表头，避免 append() 的空行问题
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="C0C0C0")
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")

            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for record in records:
            number = str(record.get("number", "") or "").strip().upper()
            if not number or number in existing_numbers:
                continue
            existing_numbers.add(number)
            row_data = [
                record.get("number", ""),
                record.get("asin", ""),
                record.get("product_url", ""),
                record.get("title", ""),
                record.get("poster_url", ""),
                record.get("search_keyword", ""),
            ]
            ws.append(row_data)

        _format_asin_worksheet(ws)

        _save_workbook_atomic(wb, excel_path)
    finally:
        wb.close()


async def save_asin_to_excel(
    records: list[AsinRecord],
    excel_path: Path | None = None,
    *,
    sheet_name: str = "ASIN 数据库",
) -> Path:
    """
    保存 ASIN 记录到 Excel 文件

    Args:
        records: ASIN 记录列表
        excel_path: Excel 文件路径，默认保存到运行目录下的 amazon_asin_database.xlsx
        sheet_name: 工作表名称

    Returns:
        Excel 文件路径

    注意：
        需要安装 openpyxl 库：pip install openpyxl
    """
    from ..models.log_buffer import LogBuffer

    try:
        import openpyxl  # noqa: F401  # 线程内还要用，提前做可用性检查给出可读错误
    except ImportError:
        LogBuffer.log().write("  ⚠️ [ASIN 数据库] 缺少 openpyxl，无法保存 amazon_asin_database.xlsx")
        raise ImportError("请安装 openpyxl 库：pip install openpyxl") from None

    if excel_path is None:
        excel_path = _get_default_excel_path()
    elif isinstance(excel_path, str):
        excel_path = Path(excel_path)

    # openpyxl 全表格式化（4 万行库秒级）挪出事件循环
    await asyncio.to_thread(_save_asin_to_excel_sync, records, excel_path, sheet_name)
    invalidate_asin_cache(excel_path)
    return excel_path


def _format_asin_worksheet(ws) -> None:
    """格式化 ASIN 数据库工作表：按番号排序 + 固定表头、自动筛选、列宽、边框、超链接、表头样式。"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        # 先按番号排序数据行（前缀字母升序 + 数字升序），再做样式
        data_rows: list[tuple] = [
            tuple(c.value for c in r[:6])
            for r in ws.iter_rows(min_row=2, max_col=6)
            if r and r[0].value is not None and str(r[0].value).strip()
        ]
        if data_rows:
            data_rows.sort(key=_asin_sort_key)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            for row in data_rows:
                ws.append(list(row))

        ws.freeze_panes = "B2"

        last_col = get_column_letter(6)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

        header_fill = openpyxl.styles.PatternFill("solid", fgColor="F2F2F2")
        header_font = openpyxl.styles.Font(bold=True, size=11)
        header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        thin = openpyxl.styles.Side(style="thin", color="D0D0D0")
        border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = border

        for row in ws.iter_rows(min_row=2, values_only=False):
            for col_idx in [2, 4]:
                cell = row[col_idx]
                val = str(cell.value or "").strip()
                if val and val.startswith("http"):
                    existing_target = cell.hyperlink.target if cell.hyperlink else None
                    if existing_target != val:
                        cell.style = "Hyperlink"
                        cell.hyperlink = val

        # 超链接处理会覆盖边框，重新设置
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = border

        # 数据行字体统一为 11pt
        data_font = openpyxl.styles.Font(size=11)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.font = data_font

        caps = {1: 20, 2: 15, 3: 50, 4: 80, 5: 50, 6: 40}
        col_max = [0] * 7
        for row in ws.iter_rows(min_row=2, values_only=True):
            for ci, cell in enumerate(row, 1):
                if cell is None or ci > 6:
                    continue
                s = str(cell)
                width = sum(2 if "\u3040" <= c <= "\u30ff" or "\u4e00" <= c <= "\u9fff" else 1 for c in s)
                col_max[ci] = max(col_max[ci], width)
        for ci in range(1, 7):
            letter = get_column_letter(ci)
            ws.column_dimensions[letter].width = min(col_max[ci] + 2, caps.get(ci, 80))
    except Exception as e:
        from ..models.log_buffer import LogBuffer

        LogBuffer.log().write(f"  ⚠️ [ASIN 数据库] 工作表格式化失败：{e}")


async def save_single_asin_record(
    number: str,
    asin: str,
    title: str = "",
    product_url: str = "",
    poster_url: str = "",
    search_keyword: str = "",
    excel_path: Path | None = None,
) -> bool:
    """
    保存单条 ASIN 记录

    Args:
        number: 影片番号
        asin: ASIN 编号（必须为 10 位字母数字）
        title: 商品标题
        product_url: 亚马逊商品详情页链接
        poster_url: 封面图片 URL
        search_keyword: 搜索关键词
        excel_path: Excel 文件路径

    Returns:
        bool: 保存成功返回 True，失败或跳过返回 False

    示例:
        success = await save_single_asin_record(
            number="ABC-123",
            asin="B0000001",
            title="作品标题",
            product_url="https://www.amazon.co.jp/dp/B0000001",
            poster_url="https://m.media-amazon.com/images/I/xxx.jpg",
        )
        if success:
            print("保存成功")
        else:
            print("保存失败或跳过")
    """
    import re

    if not asin or not asin.strip():
        return False

    asin = asin.strip().upper()

    if not re.match(r"^[A-Z0-9]{10}$", asin):
        return False

    record: AsinRecord = {
        "number": number,
        "asin": asin,
        "product_url": product_url,
        "title": title,
        "poster_url": poster_url,
        "search_keyword": search_keyword,
    }

    try:
        await save_asin_to_excel([record], excel_path)
        return True
    except Exception:
        return False


def _update_asin_record_sync(number: str, poster_url: str, excel_path: Path) -> bool:
    """update_asin_record 同步核心（to_thread 调用，防 4 万行库保存阻塞事件循环）。"""
    with _asin_db_write_lock:
        return _update_asin_record_locked(number, poster_url, excel_path)


def _update_asin_record_locked(number: str, poster_url: str, excel_path: Path) -> bool:
    try:
        import openpyxl
    except ImportError:
        return False

    if not excel_path.exists():
        return False

    wb = openpyxl.load_workbook(excel_path)
    try:
        ws = wb.active
        updated = False
        for row in ws.iter_rows(min_row=2, values_only=False):
            row_number = str(row[0].value or "").upper()
            if row_number == number.upper():
                row[4].value = poster_url
                updated = True
                break
        if updated:
            _save_workbook_atomic(wb, excel_path)
        return updated
    finally:
        wb.close()


async def update_asin_record(
    number: str,
    poster_url: str,
    excel_path: Path | None = None,
) -> bool:
    """
    更新已有 ASIN 记录的 poster_url（原地更新，不新增行）

    Args:
        number: 影片番号
        poster_url: 新的封面 URL
        excel_path: Excel 文件路径

    Returns:
        bool: 更新成功返回 True，未找到记录返回 False
    """
    if excel_path is None:
        excel_path = _get_default_excel_path()
    updated = await asyncio.to_thread(_update_asin_record_sync, number, poster_url, excel_path)
    if updated:
        invalidate_asin_cache(excel_path)
    return updated


def _replace_asin_record_sync(
    number: str,
    asin: str,
    title: str,
    product_url: str,
    poster_url: str,
    search_keyword: str,
    excel_path: Path,
) -> bool:
    """replace_asin_record 同步核心（to_thread 调用，防 4 万行库保存阻塞事件循环）。"""
    with _asin_db_write_lock:
        return _replace_asin_record_locked(number, asin, title, product_url, poster_url, search_keyword, excel_path)


def _replace_asin_record_locked(
    number: str,
    asin: str,
    title: str,
    product_url: str,
    poster_url: str,
    search_keyword: str,
    excel_path: Path,
) -> bool:
    try:
        import openpyxl
    except ImportError:
        return False

    if not excel_path.exists():
        return False

    wb = openpyxl.load_workbook(excel_path)
    try:
        ws = wb.active
        replaced = False
        for row in ws.iter_rows(min_row=2, values_only=False):
            row_number = str(row[0].value or "").upper()
            if row_number == number.upper():
                row[1].value = asin
                row[2].value = product_url
                row[3].value = title
                row[4].value = poster_url
                row[5].value = search_keyword
                replaced = True
                break
        if replaced:
            _save_workbook_atomic(wb, excel_path)
        return replaced
    finally:
        wb.close()


async def replace_asin_record(
    number: str,
    asin: str,
    title: str,
    product_url: str,
    poster_url: str,
    search_keyword: str,
    excel_path: Path | None = None,
) -> bool:
    """按番号全行替换（特典让位规则：旧记录是特典版、新记录是正品时调用）。

    保留行位置（原地替换），ASIN/标题/链接/关键词全量换新。
    """
    if excel_path is None:
        excel_path = _get_default_excel_path()
    replaced = await asyncio.to_thread(
        _replace_asin_record_sync, number, asin, title, product_url, poster_url, search_keyword, excel_path
    )
    if replaced:
        invalidate_asin_cache(excel_path)
    return replaced


async def query_asin_database(
    number: str | None = None,
    asin: str | None = None,
    excel_path: Path | None = None,
) -> list[AsinRecord]:
    """
    查询 ASIN 数据库

    Args:
        number: 按番号查询
        asin: 按 ASIN 查询
        excel_path: Excel 文件路径

    Returns:
        匹配的记录列表

    示例:
        results = await query_asin_database(number="ABC-123")
        results = await query_asin_database(asin="B0000001")
    """
    try:
        import openpyxl  # noqa: F401  # 缓存路径同样依赖 openpyxl，提前探测给出明确提示
    except ImportError:
        from ..models.log_buffer import LogBuffer

        LogBuffer.log().write("  ⚠️ [ASIN 数据库] 缺少 openpyxl，无法读取 amazon_asin_database.xlsx")
        return []

    if excel_path is None:
        excel_path = _get_default_excel_path()

    excel_path = Path(excel_path)
    index = _get_asin_index(excel_path)
    if index is None:
        return []

    if number:
        return list(index["number"].get(number.upper(), []))
    if asin:
        return list(index["asin"].get(asin.upper(), []))
    return []


async def export_asin_statistics(
    excel_path: Path | None = None,
    output_path: Path | None = None,
) -> dict:
    """
    导出 ASIN 数据库统计信息

    Returns:
        统计信息字典
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("请安装 openpyxl 库：pip install openpyxl") from None

    if excel_path is None:
        excel_path = _get_default_excel_path()

    if not excel_path.exists():
        return {}

    if output_path is None:
        output_path = excel_path.parent / "amazon_statistics.txt"

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    total_records = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue

        if len(row) < 2:
            continue

        total_records += 1

    wb.close()

    stats = {
        "total_records": total_records,
    }

    report = (
        "=" * 60
        + "\n"
        + "Amazon ASIN 数据库统计报告\n"
        + f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        + "=" * 60
        + "\n\n"
        + f"总记录数：{total_records}\n"
        + "\n"
        + "=" * 60
        + "\n"
    )
    await asyncio.to_thread(write_file_atomic, output_path, report, "utf-8")

    return stats
